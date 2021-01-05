from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import *

class Program:
    def __init__(self, program, description, value):
        self.program = program
        self.description = description
        self.value = value

def CreateAlarms(filename):

    # get workbook
    try:
        workbook = load_workbook(filename)
    except:
        return {
            'message': "Couldn't open selected file"
        }

    # get worksheet
    worksheet = workbook['Alarms']

    # build list with texts to be ignored
    ls_ignoretxts = []
    ls_ignoretxts.append('Program')
    ls_ignoretxts.append('MCP')
    ls_ignoretxts.append('R0')
    ls_ignoretxts.append('HMI')
    ls_ignoretxts.append('PFE')

    # build list with all programs
    ls_routines = []
    x = 0
    for x in range(worksheet.max_row):

        # start variables
        skip = False

        # get row
        row = x+1

        # if empty, go to next iteration
        if worksheet[f'A{row}'].value == None:
            continue

        # if xxx text found, go to next iteration
        for ignoretxt in ls_ignoretxts:
            if ignoretxt in worksheet[f'A{row}'].value:
                skip = True
        if skip:
            continue

        # if text already in list, go to next iteration
        for program in ls_routines:
            if program == worksheet[f'A{row}'].value:
                skip = True
        if skip:
            continue

        # append program to list
        ls_routines.append(worksheet[f'A{row}'].value)

    # build list with all alarms
    ls = []

    # Build list with all elements
    ls_element = []

    # BKxx
    x = 0
    for x in range(20):
        if x < 10:
            ls_element.append(f'BK0{x}')
        else:
            ls_element.append(f'BK{x}')

    # PMx
    x = 0
    for x in range(10):
        if x < 10:
            ls_element.append(f'PM{x}')

    # PPXxx
    x = 0
    for x in range(30):
        if x < 10:
            ls_element.append(f'PPX0{x}')
        else:
            ls_element.append(f'PPX{x}')

    # Cxx
    x = 0
    for x in range(60):
        if x < 10:
            ls_element.append(f'C0{x}')
        else:
            ls_element.append(f'C{x}')

    # PWSxx
    x = 0
    for x in range(20):
        ls_element.append(f'PWS{x}')

    # SBKxx
    x = 0
    for x in range(20):
        if x < 10:
            ls_element.append(f'SBK0{x}')
        else:
            ls_element.append(f'SBK{x}')

    # iterate trought all programs
    for routine in ls_routines:

        ls_obj = []
        ls_value = []
        x = 0

        # start program
        ls.append(f'########### {routine} ###########')
        ls.append('')

        # loop trought all rows
        for x in range(worksheet.max_row):

            # get row
            row = x+1

            # build list of objects
            if worksheet[f'A{row}'].value == routine:
                program = Program(worksheet[f'A{row}'].value, worksheet[f'C{row}'].value, worksheet[f'D{row}'].value)
                ls_obj.append(program)

        # loop trought all elements
        for element in ls_element:

            # get BKxx
            if 'BK' in element:
                for obj in ls_obj:

                    # fill values list
                    if f'Block {element}' in obj.description:
                        ls_value.append(obj.value)

                # if list not empty append to main list
                if ls_value:
                    # sort ascescending
                    ls_value.sort()
                    ls.append(element)

                    # append to list
                    try:
                        ls.append(f'{ls_value[0]} - {ls_value[1]}')
                    except:
                        ls.append('deu ruim tentando escrever os valores, verificar')

                    # check value numbers
                    try:
                        if not (ls_value[0] - ls_value[1] + 1 == 0):
                            ls.append('soma dos valores diferente do esperado, verificar')
                    except:
                        ls.append('deu ruim fazendo a verificação dos valores, verificar')

                    ls.append('')

                    # clear values
                    ls_value = []

            # get PMx
            if 'PM' in element:
                for obj in ls_obj:

                    # fill values list
                    if f'Manifold {element}' in obj.description:
                        ls_value.append(obj.value)

                # if list not empty append to main list
                if ls_value:
                    # sort ascescending
                    ls_value.sort()
                    ls.append(element)

                    # append to list
                    try:
                        ls.append(f'{ls_value[0]} - {ls_value[2]}')
                    except:
                        ls.append('deu ruim tentando escrever os valores, verificar')

                    # check value numbers
                    try:
                        if not (ls_value[0] - ls_value[1] + 1 + ls_value[2] == ls_value[2]):
                            ls.append('soma dos valores diferente do esperado, verificar')
                    except:
                        ls.append('deu ruim fazendo a verificação dos valores, verificar')

                    ls.append('')

                    # clear values
                    ls_value = []

            # get PPXxx
            if 'PPX' in element:
                for obj in ls_obj:

                    # fill values list
                    if f'Present {element}' in obj.description:
                        ls_value.append(obj.value)

                # if list not empty append to main list
                if ls_value:
                    # sort ascescending
                    ls_value.sort()
                    ls.append(element)

                    # append to list
                    try:
                        ls.append(f'{ls_value[0]}')
                        ls.append(f'{ls_value[1]}')
                    except:
                        ls.append('deu ruim tentando escrever os valores, verificar')

                    ls.append('')

                    # clear values
                    ls_value = []

            # get Cxx
            if 'C' in element:
                for obj in ls_obj:

                    # fill values list
                    if f'ed {element}' in obj.description:
                        ls_value.append(obj.value)

                # if list not empty append to main list
                if ls_value:
                    # sort ascescending
                    ls_value.sort()
                    ls.append(element)

                    # append to list
                    try:
                        ls.append(f'{ls_value[0]} - {ls_value[1]}')
                        ls.append(f'{ls_value[2]} - {ls_value[3]}')
                    except:
                        ls.append('deu ruim tentando escrever os valores, verificar')

                    # check value numbers
                    try:
                        if not (ls_value[0] - ls_value[1] + 1 == 0) or not (ls_value[2] - ls_value[3] + 1 == 0):
                            ls.append('soma dos valores diferente do esperado, verificar')
                    except:
                        ls.append('deu ruim fazendo a verificação dos valores, verificar')

                    ls.append('')

                    # clear values
                    ls_value = []

            # get PWSxx
            if 'PWS' in element:
                for obj in ls_obj:

                    # fill values list
                    if f'Power {element}' in obj.description:
                        ls_value.append(obj.value)

                # if list not empty append to main list
                if ls_value:
                    # sort ascescending
                    ls_value.sort()
                    ls.append(element)

                    # append to list
                    try:
                        ls.append(f'{ls_value[0]}')
                    except:
                        ls.append('deu ruim tentando escrever os valores, verificar')

                    ls.append('')

                    # clear values
                    ls_value = []

            # get SBKxx
            if 'SBK' in element:
                for obj in ls_obj:

                    # fill values list
                    if f'{element} DNET' in obj.description:
                        ls_value.append(obj.value)

                # if list not empty append to main list
                if ls_value:
                    # sort ascescending
                    ls_value.sort()
                    ls.append(element)

                    # append to list
                    try:
                        ls.append(f'{ls_value[0]} - {ls_value[2]}')
                    except:
                        ls.append('deu ruim tentando escrever os valores, verificar')

                    # check value numbers
                    try:
                        if not (ls_value[0] - ls_value[1] + 1 + ls_value[2] == ls_value[2]):
                            ls.append('soma dos valores diferente do esperado, verificar')
                    except:
                        ls.append('deu ruim fazendo a verificação dos valores, verificar')

                    ls.append('')

                    # clear values
                    ls_value = []


    # fim
    ls.append('fim')

    # save to file
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Rows can also be appended
    for row in ls:
        ws.append([row])

    # get current time
    now = datetime.now()
    now_txt = now.strftime("_%Y%m%d_%H%M%S")

    # Save the file
    wb.save(f"export{now_txt}.xlsx")

    return {
        'message': 'Exported alarms succesfully!'
    }
